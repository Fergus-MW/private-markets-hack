"""Exact-ID project slicing; originals remain immutable project artifacts."""
from datetime import datetime
import re
import io
import zipfile
from openpyxl import Workbook
from app.project_store import artifact, link


def identifier(value):
    return str(int(value)) if isinstance(value, (int, float)) else str(value).strip()


def encode(title, rows):
    book = Workbook(write_only=True)
    book.properties.created = datetime(2000, 1, 1)
    book.properties.modified = datetime(2000, 1, 1)
    sheet = book.create_sheet(title)
    for row in rows:
        sheet.append(row)
    raw = io.BytesIO()
    book.save(raw)
    output = io.BytesIO()
    with zipfile.ZipFile(raw) as source, zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as target:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, (2000, 1, 1, 0, 0, 0))
            data = source.read(name)
            if name == "docProps/core.xml":
                data = re.sub(rb"(<dcterms:modified[^>]*>)[^<]+", rb"\g<1>2000-01-01T00:00:00Z", data)
            target.writestr(info, data)
    return output.getvalue()


def prepare(store, items, fund):
    from app.workflows import workbook, CoverageError
    target = fund.get('external_ids', {}).get('corvus:legal_entity')
    if not target:
        raise CoverageError('Loader requires a Corvus legal-entity ID')
    mapping = workbook(items['mappings'][1])
    try:
        rows = mapping['LE Mapping'].iter_rows(values_only=True)
        next(rows)
        headers = next(rows)
        id_col, name_col = headers.index('Corvus LE ID'), headers.index('Legal Entity')
        names = {str(r[name_col]).strip() for r in rows if identifier(r[id_col]) == target}
    finally:
        mapping.close()
    if not names:
        raise CoverageError('No exact fund mapping in LE Mapping')
    result = dict(items)
    for role in ('draft', 'source_gl', 'reference'):
        if role not in items:
            continue
        original, content = items[role]
        book = workbook(content)
        try:
            choices = ['Investor-Level GL'] if role == 'source_gl' else ['Upload Template', 'Upload Template (VERIFIED v4c)']
            selected = [name for name in choices if name in book.sheetnames]
            if len(selected) != 1:
                raise CoverageError('Ambiguous or missing loader sheet for ' + role)
            rows = book[selected[0]].iter_rows(values_only=True)
            headers = next(rows)
            column = headers.index('Legal Entity' if role == 'source_gl' else 'Legal Entity ID')
            retained = [list(headers) + ['Original Source Row']]
            for number, row in enumerate(rows, 2):
                value = row[column]
                if (str(value).strip() in names if role == 'source_gl' else identifier(value) == target):
                    retained.append(list(row) + [number])
            if len(retained) == 1:
                raise CoverageError('No project rows in ' + role)
            title = 'Investor-Level GL' if role == 'source_gl' else ('Upload Template (VERIFIED v4c)' if role == 'reference' else 'Upload Template')
            data = encode(title, retained)
        finally:
            book.close()
        item = artifact('scoped_' + role + '.xlsx', data, derived_from=[original['key'], items['mappings'][0]['key']], role='scoped_input')
        store.bundle(artifacts=[item], links=[link(item['key'], 'derived_from', parent) for parent in item['derived_from']])
        result[role] = (item, data)
    return result
