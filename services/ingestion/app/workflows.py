"""Deterministic, restartable checking from immutable project-local artifacts."""
import csv
import hashlib
import io
import json
import os
import platform
import subprocess
import sys
import uuid
import zipfile
from collections import Counter
from datetime import date, datetime
from importlib.metadata import version
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from app.graph import key, now
from app.project_store import artifact, link

GATES = Path(__file__).parent / "gates"
ROLES = {"draft", "terms", "entity_terms", "source_gl", "mappings", "reference"}


class CoverageError(ValueError):
    pass


def quarter_bounds(quarter):
    year, number = int(quarter[:4]), int(quarter[-1])
    start = date(year, number * 3 - 2, 1)
    end = date(year + 1, 1, 1) if number == 4 else date(year, number * 3 + 1, 1)
    from datetime import timedelta
    return start, end - timedelta(days=1)


def workbook(content):
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        if sum(entry.file_size for entry in archive.infolist()) > 100 * 1024 * 1024:
            raise CoverageError("Expanded workbook exceeds 100 MiB")
    return load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)


def draft_scope(content, gate, project, fund):
    start, end = quarter_bounds(project["quarter"])
    book = workbook(content)
    try:
        if gate == "terms":
            rows = list(book["Cover"].iter_rows(values_only=True))
            cover = {str(row[0]): row[1] for row in rows[1:] if len(row) >= 2}
            if cover.get("Entity") not in [fund["name"]] + fund.get("aliases", []):
                raise CoverageError("Draft entity does not match the project's fund")
            period = cover.get("Period end")
            if isinstance(period, datetime):
                period = period.date()
            elif isinstance(period, str):
                try:
                    period = date.fromisoformat(period)
                except ValueError:
                    period = datetime.strptime(period, "%d %B %Y").date()
            if period != end:
                raise CoverageError("Draft period does not match the project's quarter")
            if "Schedule" not in book.sheetnames:
                raise CoverageError("Draft is missing Schedule")
        else:
            rows = book["Upload Template"].iter_rows(values_only=True)
            headers = next(rows)
            entity_index, date_index = headers.index("Legal Entity ID"), headers.index("GL Date")
            target = fund.get("external_ids", {}).get("corvus:legal_entity")
            if not target:
                raise CoverageError("Loader workflow requires a Corvus legal-entity ID")
            count = 0
            for row in rows:
                if not any(value is not None for value in row):
                    continue
                count += 1
                value = row[entity_index]
                identifier = str(int(value)) if isinstance(value, (int, float)) else str(value)
                period = row[date_index]
                if isinstance(period, datetime):
                    period = period.date()
                elif isinstance(period, str):
                    period = date.fromisoformat(period)
                if identifier != target or not isinstance(period, date) or not start <= period <= end:
                    raise CoverageError("Loader draft includes a different fund or quarter")
            if not count:
                raise CoverageError("Loader draft has no rows")
    except (KeyError, IndexError, StopIteration, TypeError, ValueError) as error:
        if isinstance(error, CoverageError):
            raise
        raise CoverageError("Cannot resolve draft entity and period: " + str(error)) from None
    finally:
        book.close()
    return start, end


def validate_terms(content, entity_content, as_of, fund):
    rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    if not rows:
        raise CoverageError("Terms snapshot is empty")
    seen = set()
    for row in rows:
        investor = row["investor_id"]
        if investor in seen:
            raise CoverageError("Terms snapshot has duplicate investor identities")
        seen.add(investor)
        if not row.get("source_document") or not row.get("source_clause"):
            raise CoverageError("Terms snapshot needs source document and clause citations")
        if (not row.get("valid_from") or date.fromisoformat(row["valid_from"]) > as_of
                or (row.get("valid_to") and date.fromisoformat(row["valid_to"]) < as_of)):
            raise CoverageError("Terms snapshot contains a row outside the draft date")
    entity_rows = list(csv.DictReader(io.StringIO(entity_content.decode("utf-8-sig"))))
    values = {row["term"]: row["value"] for row in entity_rows}
    if values.get("entity_id_corvus") != fund.get("external_ids", {}).get("corvus:legal_entity"):
        raise CoverageError("Entity terms do not match the project's fund identifier")
    if any(row.get("fee_basis") == "Invested Capital" for row in rows) and "invested_capital_30jun2026" not in values:
        raise CoverageError("Invested Capital terms require the invested-capital input")


def environment_fingerprint():
    return {"python": platform.python_version(), **{name: version(name) for name in ("pandas", "numpy", "openpyxl")}}


def check_coverage(store, gate, mode, inputs, ratifications):
    required = {"draft"} | ({"terms", "entity_terms"} if gate == "terms" and mode == "terms" else set())
    if gate == "loader":
        required |= {"source_gl", "mappings"}
    missing = required - inputs.keys()
    if missing:
        raise CoverageError("Missing project artifacts: " + ", ".join(sorted(missing)))
    if set(inputs) - ROLES:
        raise CoverageError("Unknown workflow input role")
    items = {}
    for role, artifact_id in inputs.items():
        try:
            items[role] = store.read_artifact(artifact_id)
        except KeyError:
            raise CoverageError("Missing project-local artifact for " + role) from None
    # Original workbooks are mandatory. Text approximations cannot drive a checker.
    for role in required & {"draft", "source_gl", "mappings"}:
        if items[role][0]["role"] != "original":
            raise CoverageError(role + " requires retained original workbook bytes; re-ingest the source")
    if gate == "terms" and mode == "terms":
        for role in ("terms", "entity_terms"):
            decision = store.get_record("decision", ratifications.get(role, "missing"))
            if not decision or decision.get("kind") != "ratification" or decision["artifact_id"] != inputs[role]:
                raise CoverageError("A named reviewer must ratify " + role + " before checking")
            if decision["sha256"] != items[role][0]["sha256"]:
                raise CoverageError("Ratification does not match artifact version")
    return items


def render_report(project, run_id, gate, as_of, inputs, results):
    lines = ["# Workflow findings", "", f"Project: {project['name']}", f"Run: {run_id}",
             f"Gate: {gate} | As of: {as_of}", "", "## Inputs", ""]
    for role, (item, _) in sorted(inputs.items()):
        lines.append(f"- {role}: {item['filename']} | SHA-256 {item['sha256']} | artifact {item['key']}")
    lines.extend(["", "## Checks", ""])
    for check in sorted(results["checks"], key=lambda c: (c.get("tier", "c"), -float(c.get("amount", 0)), c.get("check", c.get("id", "")))):
        code = check.get("id") or check.get("check")
        name = check.get("name") or check.get("check")
        lines.append(f"- **{code} — {check['status']}** (tier {check.get('tier', 'c')}): {name}")
        if check.get("detail"):
            lines.append("  " + check["detail"])
        if "observed" in check:
            lines.append(f"  Observed: {check['observed']}; expected: {check.get('expected')}")
        if check.get("amount"):
            lines.append(f"  Amount for this check: {check['amount']:,.2f} (not additive across checks)")
    lines.extend(["", "## Summary", "", json.dumps(results["summary"], sort_keys=True), ""])
    return "\n".join(lines).encode()


def run_workflow(store, gate, mode, inputs, ratifications=None):
    if gate not in {"terms", "loader"} or mode not in {"terms", "arithmetic-only", "loader"}:
        raise ValueError("Unsupported workflow")
    if (gate == "loader") != (mode == "loader"):
        raise ValueError("Mode does not match gate")
    ratifications = ratifications or {}
    manifest = store.manifest()
    if not manifest:
        raise ValueError("Materialize the project first")
    project = manifest["project"]
    gate_file = GATES / ("terms_checks.py" if gate == "terms" else "eval_loader.py")
    gate_bytes = gate_file.read_bytes()
    runtime = environment_fingerprint()
    run_id = key("workflow-v1", store.project_id, gate, mode, inputs, ratifications,
                 hashlib.sha256(gate_bytes).hexdigest(),
                 hashlib.sha256((GATES / "entity_aliases.csv").read_bytes() + Path(__file__).with_name("loader_scope.py").read_bytes()).hexdigest() if gate == "loader" else "", runtime)
    previous = store.get_record("run", run_id)
    if previous and previous["status"] == "completed":
        return previous
    token = uuid.uuid4().hex
    run = store.claim({"key": run_id, "kind": "run", "gate": gate, "mode": mode,
                       "inputs": inputs, "ratifications": ratifications, "runtime": runtime,
                       "started_at": now()}, token)
    try:
        items = check_coverage(store, gate, mode, inputs, ratifications)
        fund = store.get_record("node", project["fund_id"])
        if gate == "loader":
            from app.loader_scope import prepare
            items = prepare(store, items, fund)
        start, end = draft_scope(items["draft"][1], gate, project, fund)
        if gate == "terms" and mode == "terms":
            validate_terms(items["terms"][1], items["entity_terms"][1], end, fund)
        source_ids = sorted(set(inputs.values()))
        gate_artifact = artifact(gate_file.name, gate_bytes, derived_from=source_ids, role="checker_code")
        input_manifest = artifact("inputs.json", json.dumps({"project_id": store.project_id, "as_of": str(end),
            "gate": gate, "mode": mode, "runtime": runtime, "inputs": {
                role: {k: value[0][k] for k in ("key", "sha256", "filename")} for role, value in items.items()}}, sort_keys=True).encode(),
            derived_from=source_ids, role="input_manifest")
        store.bundle(artifacts=[gate_artifact, input_manifest], links=[
            link(item["key"], "derived_from", input_id) for item in (gate_artifact, input_manifest) for input_id in source_ids])
        if gate == "loader":
            store.bundle(artifacts=[artifact("loader_scope.py", Path(__file__).with_name("loader_scope.py").read_bytes(), derived_from=source_ids, role="checker_code")])
            store.bundle(artifacts=[artifact("entity_aliases.csv", (GATES / "entity_aliases.csv").read_bytes(), derived_from=source_ids, role="checker_rules")])
        with TemporaryDirectory(prefix="project-workflow-") as directory:
            folder = Path(directory)
            names = {}
            for role, (_, content) in items.items():
                name = role + (".csv" if role in {"terms", "entity_terms"} else ".xlsx")
                (folder / name).write_bytes(content)
                names[role] = name
            args = [sys.executable, str(gate_file.resolve())]
            if gate == "terms":
                args += ["--schedule", names["draft"], "--as-of", str(end)]
                if mode == "arithmetic-only":
                    args.append("--arithmetic-only")
                else:
                    args += ["--terms", names["terms"], "--entity-terms", names["entity_terms"]]
            else:
                args += [names["draft"], "--source", names["source_gl"], "--mappings", names["mappings"],
                         "--quarter-start", str(start), "--quarter-end", str(end)]
                if "reference" in names:
                    args += ["--key", names["reference"]]
            args += ["--json", "checks.json"]
            # Trusted bundled code only. No credentials or connector access in its environment.
            result = subprocess.run(args, cwd=folder, capture_output=True, timeout=600,
                env={"PATH": os.environ.get("PATH", "/usr/bin:/bin"), "PYTHONHASHSEED": "0", "TZ": "UTC",
                     "OPENBLAS_NUM_THREADS": "1", "OMP_NUM_THREADS": "1"})
            if result.returncode not in {0, 1} or not (folder / "checks.json").exists():
                log = artifact("checker-error.txt", result.stderr[-100000:], derived_from=source_ids, role="error_log")
                store.bundle(artifacts=[log])
                raise RuntimeError("Checker could not complete; inspect project error artifact " + log["key"])
            checks = json.loads((folder / "checks.json").read_text())
            if not isinstance(checks.get("checks"), list) or not checks["checks"]:
                raise RuntimeError("Checker returned no results")
            checks_bytes = json.dumps(checks, sort_keys=True, separators=(",", ":"), allow_nan=False).encode()
            artifacts = [artifact("checks.json", checks_bytes, derived_from=source_ids, role="check_results"),
                         artifact("checker.log", result.stdout + result.stderr, derived_from=source_ids, role="checker_log"),
                         artifact("report.md", render_report(project, run_id, gate, str(end), items, checks),
                                  derived_from=source_ids, role="report")]
        run_node = {"key": run_id, "kind": "run", "gate": gate, "mode": mode, "as_of": str(end), "inputs": inputs}
        nodes, links = [run_node], [link(run_id, "uses", item_id) for item_id in source_ids]
        links.extend(link(run_id, "uses", item["key"]) for item in (gate_artifact, input_manifest))
        links.extend(link(run_id, "produced", item["key"]) for item in artifacts)
        for index, check in enumerate(checks["checks"]):
            check_id = key(run_id, index, check)
            nodes.append({"key": check_id, "kind": "check_result", "run_id": run_id,
                          "result": check, "evidence_artifacts": source_ids})
            links.append(link(run_id, "checks", check_id))
            links.extend(link(check_id, "cites", input_id) for input_id in source_ids)
        store.bundle(nodes=nodes, artifacts=artifacts, links=links)
        return store.finish(run_id, token, "completed", {"summary": checks["summary"],
            "release_ready": not any(check["status"] in {"FAIL", "WARN", "DECISION"} for check in checks["checks"])
                             and mode != "arithmetic-only",
            "artifacts": {item["role"]: item["key"] for item in artifacts}})
    except CoverageError as error:
        return store.finish(run_id, token, "blocked", {"reason": str(error), "release_ready": False})
    except Exception as error:
        return store.finish(run_id, token, "failed", {"reason": str(error), "release_ready": False})
