"""Opt-in real workbook acceptance; fixtures stay outside the repository."""
import os
import json
import subprocess
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import MagicMock
from openpyxl import load_workbook
from app.loader_scope import prepare, encode
from app.project_store import artifact
from app.workflows import draft_scope, GATES

@unittest.skipUnless(os.environ.get('PROJECT_LOADER_FIXTURES'), 'Set PROJECT_LOADER_FIXTURES to extracted dataset 02')
class LoaderAcceptanceTests(unittest.TestCase):
    def test_reference_and_financial_fault(self):
        files=list(Path(os.environ['PROJECT_LOADER_FIXTURES']).rglob('*.xlsx'))
        items={}
        for role,prefix in [('draft','Tranche'),('mappings','Tranche'),('source_gl','Investor-Level')]:
            path=next(p for p in files if p.name.startswith(prefix)); content=path.read_bytes()
            items[role]=(artifact(path.name,content,role='original'),content)
        fund={'external_ids':{'corvus:legal_entity':'2254'}}
        scoped=prepare(MagicMock(),items,fund)
        draft_scope(scoped['draft'][1],'loader',{'quarter':'2026-Q2'},fund)
        self.assertEqual(encode('test', [['a'],[1]]),encode('test', [['a'],[1]]))
        with TemporaryDirectory() as folder:
            root=Path(folder)
            for role,(_,content) in scoped.items(): (root/(role+'.xlsx')).write_bytes(content)
            args=[sys.executable,str(GATES/'eval_loader.py'),'draft.xlsx','--source','source_gl.xlsx','--mappings','mappings.xlsx','--quarter-start','2026-04-01','--quarter-end','2026-06-30','--json','checks.json']
            def run():
                result=subprocess.run(args,cwd=root,capture_output=True,timeout=180)
                self.assertIn(result.returncode,(0,1),result.stderr.decode())
                return json.loads((root/'checks.json').read_text())
            baseline=run()
            self.assertEqual(baseline['summary'].get('FAIL',0),0)
            book=load_workbook(root/'draft.xlsx');sheet=book.active
            headers=[c.value for c in sheet[1]]
            col=headers.index('Investor Amount (Local)')+1
            sheet.cell(2,col).value=float(sheet.cell(2,col).value)+1000
            book.save(root/'draft.xlsx');book.close()
            faulty=run()
            self.assertGreater(faulty['summary'].get('FAIL',0),0)
            print('Loader baseline:',baseline['summary'],'; injected fault:',faulty['summary'])
