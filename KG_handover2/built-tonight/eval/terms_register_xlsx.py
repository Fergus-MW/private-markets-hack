"""
The fee terms register as an Excel file: the graph's terms_as_of response rendered for the administrator's own reference.
The graph is the record; this workbook is a surface of it, stamped with where it came from, and never a source. Nothing here calls a model.

Sheets: "Fee terms" (one row per investor in force on the as-of date, side-letter deviations from the LPA defaults flagged and shaded),
"Fund defaults" (entity-level LPA terms with clause references), "Changes since <date>" (cell-level diff against an earlier snapshot,
enriched with clause, effective date and delivery source from the delta file when given), "Provenance" (the graph's source ids per investor).

  uv run --with openpyxl python3 eval/terms_register_xlsx.py --terms-json graph_export/terms_as_of_2026-09-30.json \
      --previous-json graph_export/terms_as_of_2026-06-30.json --entity-terms stage0_baseline/entity_terms_v1.csv \
      --delta stage2_email/terms_delta_v1_to_v2.json --out results/fee_terms_register_2026-09-30.xlsx
  --terms-url <endpoint>  --as-of <date>   reads the live graph instead (?as_of= is appended), same as terms_checks.py
"""
import argparse, csv, hashlib, json, sys, urllib.parse, urllib.request
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ap = argparse.ArgumentParser()
ap.add_argument("--terms-json", help="saved terms_as_of response (fund_id, as_of, rows, provenance)")
ap.add_argument("--terms-url", help="live endpoint; ?as_of=<as-of> is appended")
ap.add_argument("--as-of", help="required with --terms-url")
ap.add_argument("--previous-json", help="an earlier terms_as_of response; produces the Changes sheet")
ap.add_argument("--entity-terms", help="entity_terms csv (fund-level LPA defaults)")
ap.add_argument("--delta", help="terms delta json; adds clause, effective, received and delivered-by to the Changes sheet")
ap.add_argument("--out", required=True)
a = ap.parse_args()
if not (a.terms_json or a.terms_url): ap.error("give --terms-json or --terms-url")
if a.terms_url and not a.as_of: ap.error("--as-of is required with --terms-url")

def sha(b): return hashlib.sha256(b).hexdigest()
def load_terms():
    if a.terms_url:
        url = a.terms_url + ("&" if "?" in a.terms_url else "?") + "as_of=" + urllib.parse.quote(a.as_of)
        try:
            with urllib.request.urlopen(url, timeout=20) as r: raw = r.read()
        except Exception as e: sys.exit(f"terms endpoint unreachable: {url} ({e}); fall back to --terms-json <saved response>")
        return json.loads(raw.decode("utf-8")), url, sha(raw)
    raw = Path(a.terms_json).read_bytes(); return json.loads(raw), Path(a.terms_json).name, sha(raw)
payload, source_label, source_sha = load_terms()
rows = payload["rows"]; as_of = payload.get("as_of") or a.as_of; fund_id = payload.get("fund_id", "")
prev = json.loads(Path(a.previous_json).read_text()) if a.previous_json else None
delta = json.loads(Path(a.delta).read_text()) if a.delta else None
ent = list(csv.DictReader(open(a.entity_terms))) if a.entity_terms else []
E = {r["term"]: r["value"] for r in ent}

def num(v):
    try: return float(v)
    except Exception: return None
def yn(v): return "Y" if str(v).strip().upper().startswith("Y") else "N"
def norm(v):
    n = num(v); return f"{n:.10g}" if n is not None else str(v).strip()

# fund defaults from the entity terms (None when not supplied: nothing is flagged)
D = dict(rate=num(E.get("mgmt_fee_rate_pa_default")), basis=E.get("mgmt_fee_basis_investment_period"), inside=yn(E.get("mgmt_fee_drawn_inside_commitment_default", "N")) if E else None,
         offset=num(E.get("fee_offset_pct_default")), cas=num(E.get("cas_deadline_days_default")))
def deviations(r):
    d = []
    if D["rate"] is not None and num(r["mgmt_fee_rate_pa"]) != D["rate"]: d.append("rate")
    if D["basis"] and r["fee_basis"] != D["basis"]: d.append("basis")
    if D["inside"] and yn(r["fee_inside_commitment"]) != D["inside"]: d.append("fee inside commitment")
    if D["offset"] is not None and num(r["fee_offset_pct"]) != D["offset"]: d.append("offset")
    if yn(r.get("fee_exempt", "N")) == "Y": d.append("fee exempt")
    if yn(r.get("mfn", "N")) == "Y": d.append("MFN")
    if D["cas"] is not None and num(r["cas_deadline_days"]) != D["cas"]: d.append("CAS deadline")
    return d
DEV_COL = {"rate": "mgmt_fee_rate_pa", "basis": "fee_basis", "fee inside commitment": "fee_inside_commitment", "offset": "fee_offset_pct", "fee exempt": "fee_exempt", "MFN": "mfn", "CAS deadline": "cas_deadline_days"}

INK, MUTED, GREEN, AMBER = "1C1A1A", "6B6B6B", "1E8E3E", "B26A00"
HEAD_FILL = PatternFill("solid", fgColor="EAE6E0"); DEV_FILL = PatternFill("solid", fgColor="FFE9C2"); ROW_FILL = PatternFill("solid", fgColor="FFF6E5")
thin = Side(style="thin", color="D9D4CC"); BOX = Border(bottom=thin)
COLS = [  # (header, field, width, number format)
    ("Investor id", "investor_id", 13, None), ("Investor", "investor_name", 44, None), ("Type", "investor_type", 7, None),
    ("Commitment", "commitment", 15, "#,##0"), ("Share", "commitment_share", 10, "0.0000%"),
    ("Mgmt fee rate p.a.", "mgmt_fee_rate_pa", 12, "0.00%"), ("Fee basis", "fee_basis", 17, None), ("Fee inside commitment", "fee_inside_commitment", 12, None),
    ("Offset", "fee_offset_pct", 9, "0%"), ("Fee exempt", "fee_exempt", 9, None), ("MFN", "mfn", 7, None), ("CAS deadline (days)", "cas_deadline_days", 11, "0"),
    ("Notices contact", "notices_contact", 34, None), ("Notices email", "notices_email", 30, None),
    ("Source document", "source_document", 60, None), ("Clause", "source_clause", 20, None), ("In force from", "valid_from", 13, None), ("Snapshot", "version", 9, None),
]
NUMERIC = {"commitment", "commitment_share", "mgmt_fee_rate_pa", "fee_offset_pct", "cas_deadline_days"}

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fee terms"; ws.sheet_properties.tabColor = "0B6BCB"
def put(sheet, r, c, v, bold=False, colour=None, size=None, fmt=None, fill=None, wrap=False):
    cell = sheet.cell(r, c, v); cell.font = Font(bold=bold, color=colour or INK, size=size or 10)
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if wrap: cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell
entity = E.get("entity", "") or (delta or {}).get("entity", ""); entity_id = E.get("entity_id_corvus", "")
put(ws, 1, 1, "Fee terms register", bold=True, size=14)
put(ws, 2, 1, "Rendered from the graph's terms_as_of response for the administrator's own reference. The graph is the record; this file is a copy of it on the date shown, never a source. Edit nothing by hand; re-run the script for a fresh copy.", colour=MUTED)
meta = [("Entity", f"{entity}" + (f"  ·  Corvus {entity_id}" if entity_id else "")), ("As-of", as_of), ("Rows in force", len(rows)),
        ("Source", f"{source_label}  ·  sha256 {source_sha[:8]}…" + (f"  ·  fund key {fund_id[:8]}…" if fund_id else "")),
        ("Fund defaults", (Path(a.entity_terms).name if a.entity_terms else "not supplied (no deviation flags)")),
        ("Generated", datetime.now().astimezone().isoformat(timespec="seconds"))]
for i, (k, v) in enumerate(meta, start=3): put(ws, i, 1, k, bold=True); put(ws, i, 2, v)
put(ws, 9, 1, "Shaded cells depart from the LPA defaults on the Fund defaults sheet (a side letter or an exemption). The Deviation column names the fields.", colour=AMBER)
HR = 10
for c, (h, f, w, fmt) in enumerate(COLS, start=1):
    put(ws, HR, c, h, bold=True, fill=HEAD_FILL, wrap=True).border = BOX; ws.column_dimensions[get_column_letter(c)].width = w
DEVC = len(COLS) + 1; put(ws, HR, DEVC, "Deviation from LPA", bold=True, fill=HEAD_FILL, wrap=True).border = BOX; ws.column_dimensions[get_column_letter(DEVC)].width = 34
ws.row_dimensions[HR].height = 30
rows_sorted = sorted(rows, key=lambda r: -(num(r.get("commitment")) or 0)); n_dev = 0
for i, r in enumerate(rows_sorted, start=HR + 1):
    dev = deviations(r); n_dev += bool(dev); devcols = {DEV_COL[d] for d in dev}
    for c, (h, f, w, fmt) in enumerate(COLS, start=1):
        v = r.get(f, ""); v = (num(v) if f in NUMERIC and num(v) is not None else v)
        if f == "investor_account_id" and isinstance(v, float): v = int(v)
        cell = put(ws, i, c, v, fmt=fmt, fill=(DEV_FILL if f in devcols else (ROW_FILL if dev else None)))
    put(ws, i, DEVC, ", ".join(dev) if dev else "", bold=bool(dev), colour=(AMBER if dev else None), fill=(ROW_FILL if dev else None))
LAST = HR + len(rows_sorted); TOT = LAST + 1
put(ws, TOT, 2, "Total", bold=True); put(ws, TOT, 4, f"=SUM(D{HR + 1}:D{LAST})", bold=True, fmt="#,##0"); put(ws, TOT, 5, f"=SUM(E{HR + 1}:E{LAST})", bold=True, fmt="0.0000%")
put(ws, TOT, DEVC, f'=COUNTA({get_column_letter(DEVC)}{HR + 1}:{get_column_letter(DEVC)}{LAST})&" investor(s) on non-default terms"', bold=True, colour=AMBER)
ws.freeze_panes = ws.cell(HR + 1, 3); ws.auto_filter.ref = f"A{HR}:{get_column_letter(DEVC)}{LAST}"

# ---------------------------------------------------------------- Fund defaults
fd = wb.create_sheet("Fund defaults"); fd.sheet_properties.tabColor = "8B857F"
put(fd, 1, 1, "Fund-level terms (LPA defaults every investor inherits unless a side letter says otherwise)", bold=True, size=12)
for c, (h, w) in enumerate([("Term", 40), ("Value", 44), ("Source document", 62), ("Clause", 10), ("In force from", 13)], start=1):
    put(fd, 3, c, h, bold=True, fill=HEAD_FILL).border = BOX; fd.column_dimensions[get_column_letter(c)].width = w
if ent:
    for i, r in enumerate(ent, start=4):
        v = r["value"]; n = num(v); fmt = None
        if r["term"] in ("mgmt_fee_rate_pa_default", "fee_offset_pct_default", "preferred_return", "carried_interest") and n is not None: v, fmt = n, "0.00%"
        elif r["term"] in ("total_commitments", "invested_capital_30jun2026") and n is not None: v, fmt = n, "#,##0"
        put(fd, i, 1, r["term"].replace("_", " ")); put(fd, i, 2, v, fmt=fmt); put(fd, i, 3, r["source_document"]); put(fd, i, 4, r["source_clause"]); put(fd, i, 5, r["valid_from"])
else: put(fd, 4, 1, "No entity terms file supplied.", colour=MUTED)
fd.freeze_panes = "A4"

# ---------------------------------------------------------------- Changes since the earlier snapshot
ch = wb.create_sheet(f"Changes since {prev['as_of']}" if prev else "Changes"); ch.sheet_properties.tabColor = AMBER
FIELDS = [f for _, f, _, _ in COLS if f != "version"]; LABEL = {f: h for h, f, _, _ in COLS}
def dfmt(f, v):
    n = num(v)
    if f in ("mgmt_fee_rate_pa", "fee_offset_pct") and n is not None: return f"{n * 100:.2f}%"
    if f in ("commitment",) and n is not None: return f"{n:,.0f}"
    if f == "cas_deadline_days" and n is not None: return f"{n:.0f}"
    return str(v)
put(ch, 1, 1, (f"What changed between the {prev['as_of']} and {as_of} snapshots of the register" if prev else "Changes"), bold=True, size=12)
put(ch, 2, 1, "Computed cell by cell from the two graph responses. Clause, effective date and delivery source come from the recorded delta where one exists.", colour=MUTED)
for c, (h, w) in enumerate([("Investor id", 13), ("Investor", 40), ("Field", 22), ("Was", 36), ("Now", 36), ("Clause", 9), ("Effective", 12), ("Known since", 12), ("Source document", 58), ("Delivered by", 34)], start=1):
    put(ch, 4, c, h, bold=True, fill=HEAD_FILL).border = BOX; ch.column_dimensions[get_column_letter(c)].width = w
n_changes = 0
if prev:
    P = {r["investor_id"]: r for r in prev["rows"]}; C = {r["investor_id"]: r for r in rows}; i = 5
    dclause = {c["field"]: c for c in (delta or {}).get("changes", [])} if delta else {}
    for iid in sorted(set(P) | set(C), key=lambda k: (k not in P, k not in C, k)):
        if iid not in P: put(ch, i, 1, iid); put(ch, i, 2, C[iid]["investor_name"]); put(ch, i, 3, "(investor added)"); i += 1; n_changes += 1; continue
        if iid not in C: put(ch, i, 1, iid); put(ch, i, 2, P[iid]["investor_name"]); put(ch, i, 3, "(investor no longer in force)"); i += 1; n_changes += 1; continue
        for f in FIELDS:
            if norm(P[iid].get(f, "")) == norm(C[iid].get(f, "")): continue
            use_delta = delta and delta.get("investor_id") == iid; d = dclause.get(f) if use_delta else None
            vals = [iid, C[iid]["investor_name"], LABEL[f], dfmt(f, P[iid].get(f, "")), dfmt(f, C[iid].get(f, "")),
                    (d["clause"] if d and d["clause"] != "-" else ""), (delta["effective"] if use_delta else C[iid].get("valid_from", "")),
                    (delta["received"] if use_delta else ""), (delta["source_document"] if use_delta else C[iid].get("source_document", "")), (delta["delivered_by"] if use_delta else "")]
            for c, v in enumerate(vals, start=1): put(ch, i, c, v, bold=(c == 5), fill=(DEV_FILL if c == 5 else None))
            i += 1; n_changes += 1
    if n_changes == 0: put(ch, 5, 1, "No differences between the two snapshots.", colour=GREEN)
else: put(ch, 5, 1, "No earlier snapshot supplied (--previous-json).", colour=MUTED)
ch.freeze_panes = "A5"

# ---------------------------------------------------------------- Provenance: the graph's source ids per investor
pv = wb.create_sheet("Provenance"); pv.sheet_properties.tabColor = "0B6BCB"
put(pv, 1, 1, "Source fact ids the graph holds for each investor row (its provenance block, verbatim). Use them to open the source in the graph's provenance panel.", bold=True)
for c, (h, w) in enumerate([("Investor id", 13), ("Investor", 44), ("Source ids", 70)], start=1): put(pv, 3, c, h, bold=True, fill=HEAD_FILL).border = BOX; pv.column_dimensions[get_column_letter(c)].width = w
names = {r["investor_id"]: r["investor_name"] for r in rows}
for i, (iid, ids) in enumerate(sorted((payload.get("provenance") or {}).items()), start=4):
    put(pv, i, 1, iid); put(pv, i, 2, names.get(iid, "")); put(pv, i, 3, ", ".join(ids) if isinstance(ids, list) else str(ids))
pv.freeze_panes = "A4"

Path(a.out).parent.mkdir(parents=True, exist_ok=True); wb.save(a.out)
print(f"{a.out}: {len(rows)} investors in force on {as_of}, {n_dev} on non-default terms, {n_changes} changed cells" + (f" since {prev['as_of']}" if prev else "") + f"; source {source_label} (sha256 {source_sha[:8]})")
