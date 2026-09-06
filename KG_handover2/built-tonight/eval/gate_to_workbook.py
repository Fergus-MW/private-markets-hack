"""
The Excel surface of the gate (build plan 4b, surface 2 of 3). Same run record as the page, written into a COPY of the administrator's own draft:
  - a green "QC gate" sheet in front: header with fingerprints, scoreboard (COUNTIFS/SUMIFS formulas over the findings table, not pasted numbers), findings by tier, passes shown
  - on the Schedule sheet: a "QC result" column per investor (a tick, or the failing check ids) and a cell comment on every offending cell with the check, the evidence and the source
Nothing here calls a model. The draft itself is never modified; a new file is written.

  uv run --with openpyxl python3 eval/gate_to_workbook.py --run results/runC2_q3_after_email.json --out results/q3_fee_schedule_QC.xlsx
      [--delta stage2_email/terms_delta_v1_to_v2.json]   adds the "because" line (old value, new value, clause) and the source to comments
      [--page-url <url of the rendered gate page>]
"""
import argparse, json, shutil
from pathlib import Path
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ap = argparse.ArgumentParser(); ap.add_argument("--run", required=True); ap.add_argument("--out", required=True); ap.add_argument("--delta"); ap.add_argument("--page-url", default="")
a = ap.parse_args(); HERE = Path(__file__).resolve().parent.parent
run = json.loads(Path(a.run).read_text()); delta = json.loads(Path(a.delta).read_text()) if a.delta else None
src = Path(run["schedule_file"]); src = src if src.is_absolute() else HERE / src
shutil.copyfile(src, a.out); wb = openpyxl.load_workbook(a.out); ws = wb["Schedule"]

GREEN, RED, AMBER, GREY, BLUE, INK = "1E8E3E", "C62828", "B26A00", "6B6B6B", "0B6BCB", "1C1A1A"
FILL = {"a": PatternFill("solid", fgColor="FFD9D9"), "b": PatternFill("solid", fgColor="FFE9C2"), "c": PatternFill("solid", fgColor="ECECEC")}
TIER_COLOUR = {"a": RED, "b": AMBER, "c": GREY}
CELLS = {"TC00": ["investor_id"], "TC01": ["rate_applied"], "TC02": ["fee_basis_applied", "basis_amount"], "TC03": ["fee_inside_commitment_applied", "unfunded_end"], "TC04": ["net_fee"],
         "TC05": ["offset_pct_applied"], "TC06": ["commitment_share"], "TC07": ["gross_fee"], "TC09": ["net_fee"], "TC10": ["called_end", "unfunded_end"]}
FIELDS = {"TC01": ["mgmt_fee_rate_pa"], "TC02": ["fee_basis"], "TC05": ["fee_offset_pct"], "TC09": ["mgmt_fee_rate_pa", "fee_basis", "fee_offset_pct"], "TC03": ["fee_inside_commitment"]}
LABEL = {"mgmt_fee_rate_pa": "rate", "fee_basis": "basis", "fee_offset_pct": "offset", "fee_inside_commitment": "fee inside commitment"}
def fmt(field, v):
    if field in ("mgmt_fee_rate_pa", "fee_offset_pct"):
        try: return f"{float(v) * 100:.2f}%"
        except Exception: return str(v)
    return str(v)
def because(check, investor_id):
    if not delta or delta.get("investor_id") != investor_id: return "", ""
    parts = [f"{LABEL.get(c['field'], c['field'])} {fmt(c['field'], c['new'])} (was {fmt(c['field'], c['old'])}, cl. {c['clause']})" for c in delta["changes"] if c["field"] in FIELDS.get(check, [])]
    source = f"{delta['source_document']} · in force from {delta['effective']} · known since {delta['received']} · {delta['delivered_by']}"
    return " · ".join(parts), source

# ---------------------------------------------------------------- schedule sheet: rows, columns, comments, result column
head = {c.value: c.column for c in ws[1] if c.value}; rows = {}
for r in range(2, ws.max_row + 1):
    iid = ws.cell(r, head["investor_id"]).value; name = ws.cell(r, head["investor_name"]).value
    if iid is not None: rows[str(name)] = (r, str(iid))
res_col = ws.max_column + 1; ws.cell(1, res_col, "QC result").font = Font(bold=True); ws.cell(1, res_col + 1, "QC evidence").font = Font(bold=True)
per_row = {}   # row -> list of (check, tier, evidence)
findings = [c for c in run["checks"] if c["status"] in ("FAIL", "WARN", "DECISION")]
for c in findings:
    evid = {seg.split(": ", 1)[0]: seg.split(": ", 1)[1] for seg in c["detail"].split("; ") if ": " in seg} if c.get("detail") else {}
    names = [n for n in c["investors"].split(", ") if n] if c.get("investors") else []
    if c["check"] == "TC08":  # totals row
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, head["investor_id"]).value == "TOTAL":
                cell = ws.cell(r, head["net_fee"]); cell.fill = FILL[c["tier"]]; cell.comment = Comment(f"{c['check']} (tier {c['tier']}): {c['name']}\n{c.get('detail', '')}", "QC gate"); per_row.setdefault(r, []).append((c["check"], c["tier"], c.get("detail", "")))
        continue
    for n in names:
        if n not in rows: continue
        r, iid = rows[n]; ev = evid.get(n, ""); why, source = because(c["check"], iid)
        text = f"{c['check']} (tier {c['tier']}): {c['name']}\nevidence: {ev}" + (f"\nbecause: {why}" if why else "") + (f"\nsource: {source}" if source else "")
        for col in CELLS.get(c["check"], []):
            if col in head:
                cell = ws.cell(r, head[col]); cell.fill = FILL[c["tier"]]
                cell.comment = Comment(text if cell.comment is None else cell.comment.text + "\n\n" + text, "QC gate"); cell.comment.width, cell.comment.height = 420, 180
        per_row.setdefault(r, []).append((c["check"], c["tier"], ev))
for name, (r, iid) in rows.items():
    hits = per_row.get(r, [])
    if hits:
        worst = min(h[1] for h in hits); ws.cell(r, res_col, "✗ " + ", ".join(sorted({h[0] for h in hits}))).font = Font(bold=True, color=TIER_COLOUR[worst])
        ws.cell(r, res_col + 1, " | ".join(h[2] for h in hits if h[2]))
    else:
        ws.cell(r, res_col, "✓").font = Font(bold=True, color=GREEN)
for r in range(2, ws.max_row + 1):
    if ws.cell(r, head["investor_id"]).value == "TOTAL" and r not in per_row: ws.cell(r, res_col, "✓").font = Font(bold=True, color=GREEN)
ws.column_dimensions[get_column_letter(res_col)].width = 16; ws.column_dimensions[get_column_letter(res_col + 1)].width = 60

# ---------------------------------------------------------------- QC gate sheet, in front, green tab
q = wb.create_sheet("QC gate", 0); q.sheet_properties.tabColor = GREEN
def put(r, c, v, bold=False, colour=None, size=None):
    cell = q.cell(r, c, v); cell.font = Font(bold=bold, color=colour or INK, size=size or 11); return cell
put(1, 1, "QC gate result", True, size=16); put(2, 1, "Generated from the gate's run record; the draft itself is untouched. Passes are shown, never hidden. A decision owed is not an error.", colour=GREY)
hdr = [("Draft", Path(run["schedule_file"]).name + f"  ·  sha256 {(run['schedule_sha256'] or '')[:8]}…"), ("Entity", f"{run['entity']}  ·  Corvus {run.get('entity_id') or '?'}"), ("As-of", run["as_of"]),
       ("Terms snapshot", (Path(str(run["terms_file"])).name if run.get("terms_file") else "none (arithmetic only)") + (f"  ·  sha256 {run['terms_snapshot_sha256'][:8]}…  ·  {run['terms_rows_in_force']} rows in force" if run.get("terms_snapshot_sha256") else "")),
       ("Run", f"{run['run_id']}  ·  {run['run_at']}  ·  mode {run['mode']}"), ("Page", a.page_url or "")]
for i, (k, v) in enumerate(hdr): put(4 + i, 1, k, True); put(4 + i, 2, v)
# findings table (formulas below point at it)
F0 = 20; put(F0 - 1, 1, "Findings", True, size=13)
cols = ["Check", "Tier", "Status", "Investor", "Amount", "Evidence", "Because", "Source"]
for j, h in enumerate(cols): put(F0, 1 + j, h, True)
ordered = sorted(findings, key=lambda c: ({"a": 0, "b": 1, "c": 2}.get(c["tier"], 3), -float(c.get("amount") or 0)))
r = F0 + 1
for c in ordered:
    names = [n for n in c["investors"].split(", ") if n] if c.get("investors") else [""]
    evid = {seg.split(": ", 1)[0]: seg.split(": ", 1)[1] for seg in c["detail"].split("; ") if ": " in seg} if c.get("detail") else {}
    for n in names:
        iid = rows.get(n, (None, ""))[1]; why, source = because(c["check"], iid)
        vals = [c["check"], c["tier"], c["status"], n, float(c.get("amount") or 0), evid.get(n, c.get("detail", "")), why, source]
        for j, v in enumerate(vals): put(r, 1 + j, v, colour=TIER_COLOUR.get(c["tier"]) if j in (0, 1, 2) else None)
        q.cell(r, 5).number_format = "#,##0.00"; r += 1
F1 = max(r - 1, F0 + 1)
P0 = F1 + 3; put(P0 - 1, 1, "Passes", True, size=13)
for j, h in enumerate(["Check", "Tier", "Name"]): put(P0, 1 + j, h, True)
r = P0 + 1
for c in run["checks"]:
    if c["status"] == "PASS": put(r, 1, "✓ " + c["check"], colour=GREEN); put(r, 2, c["tier"]); put(r, 3, c["name"]); r += 1
    elif c["status"] == "SKIPPED": put(r, 1, "– " + c["check"], colour=GREY); put(r, 2, c["tier"]); put(r, 3, c["name"], colour=GREY); r += 1
# scoreboard as formulas over the findings table (client rule: no pasted numbers in a control tab)
tier_rng = f"$B${F0 + 1}:$B${F1}"; st_rng = f"$C${F0 + 1}:$C${F1}"; amt_rng = f"$E${F0 + 1}:$E${F1}"
S0 = 11; put(S0, 1, "Scoreboard", True, size=13)
board = [("Tier a", f'=COUNTIFS({tier_rng},"a",{st_rng},"FAIL")', RED), ("Tier b", f'=COUNTIFS({tier_rng},"b",{st_rng},"FAIL")', AMBER), ("Tier c", f'=COUNTIFS({tier_rng},"c",{st_rng},"FAIL")', GREY),
         ("Decisions owed", f'=COUNTIF({st_rng},"DECISION")', BLUE), ("Passes", f'=COUNTIF($A${P0 + 1}:$A${r},"✓*")&" of {len(run["checks"])}"', GREEN), ("Amount at stake (tier a)", f'=SUMIFS({amt_rng},{tier_rng},"a",{st_rng},"FAIL")', INK)]
for j, (k, f, colour) in enumerate(board): put(S0 + 1, 1 + j, k, True, colour=GREY); cell = put(S0 + 2, 1 + j, f, True, colour=colour, size=14); cell.number_format = "#,##0.00" if "SUMIFS" in f else "General"
put(S0 + 3, 1, "Every number above is a formula over the findings table below; change nothing by hand.", colour=GREY)
for col, w in zip("ABCDEFGH", (26, 10, 12, 34, 14, 60, 60, 70)): q.column_dimensions[col].width = w
for row in q.iter_rows(min_row=F0, max_row=r): 
    for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
q.freeze_panes = "A4"
wb.save(a.out)
print(f"written {a.out}: {len(findings)} finding rows, {sum(1 for c in run['checks'] if c['status'] == 'PASS')} passes, comments on {sum(1 for row in ws.iter_rows() for cell in row if cell.comment)} cells")
